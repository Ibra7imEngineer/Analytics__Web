"""
Universal Data Extraction Portal - Professional Edition
=======================================================
Extract text, tables, and data from URLs and files with professional UI and exports.

Author: BLACKBOXAI Enhanced
Version: 2.0.0 - Bugfixed & Professional
License: MIT
"""

# ============================================================================
# IMPORTS
# ============================================================================
import io
import json
import re
import subprocess
import sys
import time
from typing import List, Tuple

import pandas as pd
import streamlit as st
from bs4 import BeautifulSoup
import requests
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError, sync_playwright

# OPTIONAL for Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = Font = PatternFill = Alignment = None

# CONSTANTS
PAGE_LOAD_TIMEOUT = 60_000  # milliseconds for Playwright load state waits
SMART_ELEMENT_FAST_WAIT = 2_000
SMART_ELEMENT_EXTENDED_WAIT = 15_000
SMART_ELEMENT_MAX_WAIT = 60_000
HEAVY_SPA_SCRIPT_THRESHOLD = 20
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB
SUPPORTED_EXTENSIONS = ['.csv', '.xlsx', '.xls', '.json', '.txt', '.html']

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================
def clean_text(text: str) -> str:
    """Clean text content."""
    if not text:
        return ""
    text = re.sub(r'\s+', ' ', text.strip())
    return ''.join(char for char in text if char.isprintable())

def validate_url(url: str) -> bool:
    """Validate URL."""
    try:
        result = requests.utils.urlparse(url)
        return all([result.scheme, result.netloc])
    except:
        return False

def extract_tables_from_html(html_content: str) -> List[pd.DataFrame]:
    """Extract tables from HTML."""
    if not html_content:
        return []
    try:
        soup = BeautifulSoup(html_content, 'html.parser')
        tables = soup.find_all('table')
        dataframes = []
        for table in tables:
            try:
                df = pd.read_html(str(table))[0]
                if not df.empty:
                    dataframes.append(df)
            except:
                continue
        return dataframes
    except:
        return []

# ============================================================================
# ENTRY PARSING HELPERS
# ============================================================================

def remove_noise_elements(soup: BeautifulSoup):
    """Remove scripts, ads, and noisy page elements."""
    for tag in soup(["script", "style", "noscript", "iframe", "ins", "header", "footer", "nav", "aside"]):
        tag.decompose()

    noise_keywords = [
        'ad', 'ads', 'banner', 'promo', 'cookie', 'subscribe', 'welcome', 'modal', 'popup',
        'recommended', 'advertisement', 'announcement', 'offer', 'sponsor'
    ]
    for keyword in noise_keywords:
        pattern = re.compile(fr'(^|[^A-Za-z0-9]){re.escape(keyword)}([^A-Za-z0-9]|$)', re.I)
        for node in soup.find_all(True, {'class': pattern}):
            node.decompose()
        for node in soup.find_all(True, {'id': pattern}):
            node.decompose()


def get_clean_text(element) -> str:
    return clean_text(element.get_text(separator=' ', strip=True))


def extract_metadata_values(text: str) -> dict:
    """Extract structured metadata fields from text."""
    metadata = {
        'year': None,
        'item_type': None,
        'rating': None,
        'duration': None,
        'extra': None,
    }
    if not text:
        return metadata

    normalized = re.sub(r'[\s]+', ' ', text)

    year_match = re.search(r'\b(19|20)\d{2}\b', normalized)
    if year_match:
        metadata['year'] = year_match.group(0)

    rating_match = re.search(r'(?:تقييم|rating|تصنيف)\s*[:\-]?\s*(\d+(?:\.\d+)?)(?:\s*/\s*10)?', normalized, re.I)
    if rating_match:
        metadata['rating'] = rating_match.group(1)
    else:
        slash_match = re.search(r'\b(\d+(?:\.\d+)?)/10\b', normalized)
        if slash_match:
            metadata['rating'] = slash_match.group(1)
        else:
            star_match = re.search(r'([★☆]{2,10})', normalized)
            if star_match:
                metadata['rating'] = str(len(star_match.group(1)))

    duration_match = re.search(r'\b(\d{1,2}\s*(?:ساعة|س|دقيقة|دقائق|hr|h|min|m))\b', normalized, re.I)
    if duration_match:
        metadata['duration'] = duration_match.group(1)
    else:
        time_match = re.search(r'\b(\d{1,2}:\d{2})\b', normalized)
        if time_match:
            metadata['duration'] = time_match.group(1)

    item_types = ['فيلم', 'مسلسل', 'لاعب', 'مباراة', 'دوري', 'برنامج', 'ألبوم', 'حفلة', 'فنان', 'مسابقة', 'ترتيب']
    for item in item_types:
        if re.search(fr'\b{re.escape(item)}\b', normalized, re.I):
            metadata['item_type'] = item
            break

    metadata['extra'] = normalized
    return metadata


def get_entry_primary_title(entry: BeautifulSoup) -> str:
    """Extract the main title from an entry container."""
    candidates = []
    for selector in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6', '.title', '.name', '.heading', 'strong', 'b', 'a']:
        for node in entry.select(selector):
            text = get_clean_text(node)
            if text and len(text) > 2 and not re.search(r'\b(مشاهدة|تحميل|اضغط|تفاصيل|عرض|More|Read More)\b', text, re.I):
                candidates.append(text)
        if candidates:
            return candidates[0]
    fallback = get_clean_text(entry)
    return fallback.split(' | ')[0].split('\n')[0].strip() if fallback else 'Unknown'


def get_entry_media_link(entry: BeautifulSoup, base_url: str = None) -> str:
    """Extract the best image or poster link from an entry."""
    img = entry.find('img')
    if img and img.get('src'):
        return requests.compat.urljoin(base_url or '', img.get('src'))
    if img and img.get('data-src'):
        return requests.compat.urljoin(base_url or '', img.get('data-src'))
    return ''


def get_entry_source_url(entry: BeautifulSoup, base_url: str = None) -> str:
    """Extract the detail page URL from an entry."""
    for link in entry.find_all('a', href=True):
        href = link.get('href')
        if href and not href.startswith('#') and not href.lower().startswith('javascript:'):
            return requests.compat.urljoin(base_url or '', href)
    return ''


def parse_entry(entry: BeautifulSoup, base_url: str = None) -> dict:
    """Parse a single entry container into a clean data object."""
    title = get_entry_primary_title(entry)
    metadata_texts = []
    for node in entry.find_all(['span', 'div', 'p', 'li']):
        text = get_clean_text(node)
        if text and 2 < len(text) <= 80:
            metadata_texts.append(text)
    metadata = extract_metadata_values(' | '.join(metadata_texts))
    return {
        'primary_title': title,
        'year': metadata['year'] or '',
        'item_type': metadata['item_type'] or '',
        'rating': metadata['rating'] or '',
        'duration': metadata['duration'] or '',
        'media_link': get_entry_media_link(entry, base_url),
        'source_url': get_entry_source_url(entry, base_url),
        'raw_metadata': metadata['extra'] or ''
    }


def extract_entries_from_soup(soup: BeautifulSoup, base_url: str = None) -> List[dict]:
    """Find repeating entry containers and extract structured records."""
    remove_noise_elements(soup)
    candidates = {}
    for node in soup.find_all(['article', 'li', 'div', 'section', 'tr']):
        if node.name == 'body':
            continue
        text = get_clean_text(node)
        if not text or len(text) < 30:
            continue
        cls = ' '.join(node.get('class') or [])
        signature = f'{node.name}|{cls}'
        candidates.setdefault(signature, []).append(node)

    repeated_groups = [group for group in candidates.values() if len(group) > 2]
    repeated_groups.sort(key=lambda group: (-len(group), len(get_clean_text(group[0]))))

    entries = []
    if repeated_groups:
        top_group = repeated_groups[0]
        for node in top_group:
            if any(parent in top_group for parent in node.parents):
                continue
            entries.append(parse_entry(node, base_url))
    else:
        fallback = soup.find_all(['article', 'section', 'div'], limit=20)
        for node in fallback:
            if get_clean_text(node) and node.find(['h1', 'h2', 'h3', 'a']):
                entries.append(parse_entry(node, base_url))
                if len(entries) >= 10:
                    break

    cleaned_entries = []
    seen = set()
    for entry in entries:
        key = (entry['primary_title'], entry['source_url'], entry['media_link'])
        if key not in seen and entry['primary_title'] != 'Unknown':
            seen.add(key)
            cleaned_entries.append(entry)
    return cleaned_entries

# ============================================================================
# CORE EXTRACTION
# ============================================================================

def is_heavy_spa(page) -> bool:
    """Detect heavy SPA pages and adjust buffer timing automatically."""
    try:
        root_app = page.query_selector("div#root, div#app, div#__next, div[data-reactroot], div[data-vue-root]")
        many_scripts = len(page.query_selector_all("script")) >= HEAVY_SPA_SCRIPT_THRESHOLD
        framework_flag = page.evaluate(
            "() => !!(window.__NEXT_DATA__ || window.__NUXT__ || window.__REACT_DATA__ || window.Vue || window.angular || window.__SVELTE__)")
        return bool(root_app and (many_scripts or framework_flag))
    except Exception:
        return False


def smart_wait_for_data_element(page):
    """Wait for a data-containing element using progressive polling and safe time bounds."""
    selectors = [
        "div[data-testid='content']",
        "div[class*='content']",
        "div[class*='main']",
        "main",
        "article",
        "section",
        "table",
        "div[role='main']",
    ]

    for selector in selectors:
        try:
            page.wait_for_selector(selector, state="visible", timeout=SMART_ELEMENT_FAST_WAIT)
            return selector
        except PlaywrightTimeoutError:
            continue

    for selector in selectors:
        try:
            page.wait_for_selector(selector, state="visible", timeout=SMART_ELEMENT_EXTENDED_WAIT)
            return selector
        except PlaywrightTimeoutError:
            continue

    deadline = time.monotonic() + SMART_ELEMENT_MAX_WAIT / 1000
    poll_interval = 1.0
    while time.monotonic() < deadline:
        for selector in selectors:
            if page.query_selector(selector):
                return selector
        time.sleep(poll_interval)
        if poll_interval < 1.0:
            poll_interval += 0.25
    return None


def scrape_url_requests(url: str, timeout: int = 15) -> Tuple[str, List[pd.DataFrame], List[dict]]:
    """Fallback scraping with requests for environments without Playwright browser support."""
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    try:
        response = requests.get(url, headers=headers, timeout=timeout)
        response.raise_for_status()
        html_content = response.content.decode('utf-8', errors='ignore')
        soup = BeautifulSoup(html_content, 'html.parser')
        remove_noise_elements(soup)
        text_content = get_clean_text(soup)
        tables = extract_tables_from_html(html_content)
        entries = extract_entries_from_soup(soup, base_url=url)
        return text_content or "No text found.", tables, entries
    except Exception as e:
        raise RuntimeError(f"HTTP fallback scrape failed: {e}") from e


def install_playwright_browsers() -> None:
    """Install Playwright browsers when missing."""
    st.warning("Playwright browser not found. Attempting to install browsers automatically...")
    try:
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium", "firefox", "webkit"],
            check=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
        )
        st.success("Playwright browser installation completed successfully.")
    except subprocess.CalledProcessError as install_error:
        raise RuntimeError(
            f"Failed to install Playwright browsers: {install_error.stderr.strip() or install_error}") from install_error
    except Exception as install_error:
        raise RuntimeError(
            f"Could not run Playwright installer: {install_error}") from install_error


def try_launch_playwright_browser(p):
    """Try launching available Playwright browsers in order."""
    engines = ["chromium", "firefox", "webkit"]
    last_error = None
    for engine in engines:
        try:
            browser = getattr(p, engine).launch(headless=True)
            st.info(f"Using Playwright engine: {engine}")
            return browser
        except Exception as launch_error:
            last_error = launch_error
            continue
    if last_error:
        raise last_error


def scrape_url(url: str) -> Tuple[str, List[pd.DataFrame], List[dict]]:
    """Scrape URL with Playwright-based intelligent waiting."""
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
    try:
        with sync_playwright() as p:
            try:
                browser = try_launch_playwright_browser(p)
            except Exception as launch_error:
                error_text = str(launch_error).lower()
                if "executable doesn't exist" in error_text or "playwright install" in error_text or "libglib-2.0.so.0" in error_text:
                    try:
                        install_playwright_browsers()
                        browser = try_launch_playwright_browser(p)
                    except Exception:
                        st.warning("Playwright browser engines are unavailable in this environment; falling back to HTTP scraping.")
                        return scrape_url_requests(url)
                else:
                    st.warning("Playwright browser launch failed; falling back to HTTP scraping.")
                    return scrape_url_requests(url)

            page = browser.new_page(user_agent=headers['User-Agent'])
            page.goto(url, timeout=PAGE_LOAD_TIMEOUT, wait_until='domcontentloaded')
            page.wait_for_load_state('networkidle', timeout=PAGE_LOAD_TIMEOUT)

            if is_heavy_spa(page):
                page.wait_for_timeout(3_000)

            smart_wait_for_data_element(page)
            html_content = page.content()

        soup = BeautifulSoup(html_content, 'html.parser')
        remove_noise_elements(soup)
        text_content = get_clean_text(soup)
        tables = extract_tables_from_html(html_content)
        entries = extract_entries_from_soup(soup, base_url=url)
        return text_content or "No text found.", tables, entries
    except PlaywrightTimeoutError as e:
        raise RuntimeError("Timeout while waiting for the page to become idle or present data elements.") from e
    except Exception as e:
        error_text = str(e).lower()
        fallback_keywords = [
            "target page, context or browser has been closed",
            "browser has been closed",
            "closed",
            "libglib-2.0.so.0",
            "cannot open shared object file",
            "executable doesn't exist",
            "playwright install",
            "failed to launch",
        ]
        if any(keyword in error_text for keyword in fallback_keywords):
            st.warning("Playwright browser startup failed in this environment; falling back to HTTP scraping.")
            return scrape_url_requests(url)

        raise RuntimeError(
            f"Playwright scrape failed: {e}. Ensure Playwright is installed and browsers are available via `playwright install chromium`.") from e

def process_file(uploaded_file) -> Tuple[str, List[pd.DataFrame], List[dict]]:
    """Process uploaded file."""
    file_name = uploaded_file.name.lower()
    try:
        if file_name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
            return df.to_string(index=False), [df], []
        elif file_name.endswith(('.xlsx', '.xls')):
            df_dict = pd.read_excel(uploaded_file, sheet_name=None)
            if isinstance(df_dict, dict):
                text_parts = [f"Sheet {k}:\n{v.to_string(index=False)}" for k, v in df_dict.items()]
                return '\n\n'.join(text_parts), list(df_dict.values()), []
            text_content = df_dict.to_string(index=False)
            return text_content, [df_dict], []
        elif file_name.endswith('.json'):
            data = json.load(uploaded_file)
            return json.dumps(data, indent=2, ensure_ascii=False), [], []
        else:
            content = uploaded_file.read().decode('utf-8')
            if file_name.endswith('.html'):
                soup = BeautifulSoup(content, 'html.parser')
                remove_noise_elements(soup)
                text = get_clean_text(soup)
                tables = extract_tables_from_html(content)
                entries = extract_entries_from_soup(BeautifulSoup(content, 'html.parser'))
                return text, tables, entries
            return clean_text(content), [], []
    except Exception as e:
        raise RuntimeError(f"File process failed: {e}")

# ============================================================================
# EXPORT FUNCTIONS
# ============================================================================
def generate_professional_excel(source_name: str, text_content: str, tables: List[pd.DataFrame]) -> bytes:
    """Generate styled Excel."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("Install openpyxl for Excel export.")
    wb = Workbook()
    wb.remove(wb.active)

    # Text sheet
    if text_content:
        ws_text = wb.create_sheet("Extracted_Text")
        ws_text['A1'] = f"Source: {source_name}"
        ws_text['A1'].font = Font(bold=True, size=14)
        ws_text['A2'] = text_content[:30000]

    # Tables
    for i, df in enumerate(tables, 1):
        ws = wb.create_sheet(f"Table_{i}")
        # Headers
        for col_num, header in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=str(header))
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="4F94CD")
            cell.alignment = Alignment(horizontal="center")
        # Data
        for row_num, (_, row) in enumerate(df.iterrows(), 2):
            for col_num, value in enumerate(row.values, 1):
                ws.cell(row=row_num, column=col_num, value=str(value))
        # Auto width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def generate_entries_excel(source_name: str, df: pd.DataFrame) -> bytes:
    """Generate Excel workbook for extracted entries."""
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("Install openpyxl for Excel export.")
    wb = Workbook()
    ws = wb.active
    ws.title = "Entries"

    headers = list(df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=str(header))
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4F94CD")
        cell.alignment = Alignment(horizontal="center")

    for row_num, (_, row) in enumerate(df.iterrows(), 2):
        for col_num, value in enumerate(row.values, 1):
            ws.cell(row=row_num, column=col_num, value=str(value))

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()

# ============================================================================
# UI FUNCTIONS
# ============================================================================
@st.cache_data
def setup_page():
    st.set_page_config(
        page_title="DATAX - Enterprise Data Extraction",
        page_icon="🔍",
        layout="wide"
    )
    # Dark Mode Glassmorphism CSS
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=IBM+Plex+Sans+Arabic:wght@300;400;500;600;700&display=swap');
    html, body, [class*="css"]  {
        font-family: 'Inter', 'IBM Plex Sans Arabic', sans-serif;
        direction: rtl;
        color: #e5e7eb;
    }
    body {
        background: #10131f;
    }
    .stApp {
        background: transparent;
        animation: fadeInSmooth 0.8s ease both;
    }
    .main {
        background: linear-gradient(135deg, #0f1425 0%, #111828 45%, #141b2d 100%);
        padding-top: 1rem;
        min-height: 100vh;
    }
    .glass-card {
        background: rgba(17, 25, 40, 0.75);
        backdrop-filter: blur(15px);
        -webkit-backdrop-filter: blur(15px);
        border: 1px solid rgba(255,255,255,0.05);
        border-radius: 28px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 18px 60px rgba(0,0,0,0.35);
        transition: transform 0.4s ease, opacity 0.4s ease;
    }
    .glass-card:hover {
        transform: translateY(-1px);
    }
    .result-panel {
        padding: 1rem 1rem 1.5rem 1rem;
    }
    .breadcrumb {
        display: inline-flex;
        flex-wrap: wrap;
        gap: 0.4rem;
        align-items: center;
        color: #9ca3af;
        font-size: 0.95rem;
        margin-bottom: 0.75rem;
    }
    .breadcrumb span {
        color: #d1d5db;
    }
    .toast-notice {
        display: inline-flex;
        align-items: center;
        gap: 0.5rem;
        padding: 0.65rem 1rem;
        border-radius: 999px;
        background: rgba(16, 185, 129, 0.12);
        border: 1px solid rgba(16, 185, 129, 0.25);
        color: #a7f3d0;
        font-size: 0.95rem;
        margin-bottom: 1.25rem;
        backdrop-filter: blur(10px);
    }
    .metrics-grid {
        display: grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: 1rem;
        margin-bottom: 1.5rem;
    }
    .metric-card {
        background: rgba(8, 14, 26, 0.9);
        border: 1px solid rgba(56, 189, 248, 0.18);
        border-left: 4px solid rgba(34, 211, 238, 0.9);
        border-radius: 18px;
        padding: 1.3rem 1.2rem;
        color: #e5e7eb;
        box-shadow: 0 10px 30px rgba(0,0,0,0.24);
    }
    .metric-card h3 {
        margin: 0;
        font-size: 0.95rem;
        color: #9ca3af;
        letter-spacing: 0.04em;
    }
    .metric-card .metric-value {
        font-size: 2.3rem;
        font-weight: 700;
        letter-spacing: 0.04em;
        line-height: 1.05;
        margin-top: 0.7rem;
        color: #ffffff;
        font-family: 'Inter', sans-serif;
    }
    .metric-accent-cyan { border-left-color: rgba(34, 211, 238, 0.95); }
    .metric-accent-emerald { border-left-color: rgba(16, 185, 129, 0.95); }
    .stTabs [data-baseweb="tab-list"] {
        background: transparent;
        border: none;
        display: flex;
        gap: 0.8rem;
        padding: 0;
        margin-bottom: 1rem;
    }
    .stTabs [data-baseweb="tab"] {
        flex: 1;
        min-width: 10rem;
        background: rgba(15,23,42,0.88);
        border: 1px solid rgba(148,163,184,0.18);
        border-radius: 20px;
        padding: 0.85rem 1rem;
        color: #d1d5db;
        box-shadow: 0 10px 24px rgba(0,0,0,0.18);
        transition: transform 0.2s ease, background 0.2s ease, border-color 0.2s ease, color 0.2s ease;
        font-size: 0.95rem;
        font-weight: 500;
    }
    .stTabs [data-baseweb="tab"]:hover {
        transform: translateY(-1px);
        background: rgba(30,41,59,0.98);
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background: linear-gradient(135deg, rgba(34,211,238,0.16), rgba(16,185,129,0.16));
        border-color: rgba(34,211,238,0.44);
        color: #ffffff;
        box-shadow: 0 16px 36px rgba(34,211,238,0.14);
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] span {
        font-weight: 700;
    }
    .action-bar {
        display: flex;
        justify-content: space-between;
        align-items: center;
        gap: 1rem;
        flex-wrap: wrap;
        margin-bottom: 1.25rem;
    }
    .action-toolbar {
        display: flex;
        align-items: center;
        gap: 0.75rem;
    }
    .icon-button button, .icon-button input {
        border: 1px solid rgba(255,255,255,0.12);
        background: rgba(255,255,255,0.04);
        color: #f8fafc;
        border-radius: 999px;
        width: 3.4rem;
        height: 3.4rem;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        font-size: 1.2rem;
        box-shadow: 0 12px 22px rgba(0,0,0,0.2);
        transition: transform 0.2s ease, background 0.2s ease, border-color 0.2s ease;
    }
    .icon-button button:hover, .icon-button input:hover {
        transform: translateY(-1px);
        background: rgba(255,255,255,0.12);
        border-color: rgba(34, 211, 238, 0.35);
    }
    .search-filter input {
        background: rgba(255,255,255,0.06) !important;
        border: 1px solid rgba(148, 163, 184, 0.22) !important;
        color: #f8fafc !important;
        border-radius: 14px !important;
        padding: 0.85rem 1rem !important;
        min-width: 320px;
    }
    .dataframe-container {
        overflow: hidden;
        border-radius: 20px;
        border: 1px solid rgba(255,255,255,0.08);
    }
    .stDataFrame > div > div > div {
        background: transparent !important;
    }
    .stDataFrame table {
        border: none !important;
        background: rgba(15,23,42,0.88);
    }
    .stDataFrame thead tr {
        position: sticky;
        top: 0;
        background: rgba(30, 41, 59, 0.95);
        color: #e2e8f0;
        z-index: 2;
    }
    .stDataFrame th, .stDataFrame td {
        border: none !important;
        padding: 0.75rem 0.9rem !important;
    }
    .stDataFrame tbody tr:nth-child(odd) {
        background: rgba(255,255,255,0.02);
    }
    .stDataFrame tbody tr:nth-child(even) {
        background: rgba(255,255,255,0.01);
    }
    .stDataFrame tbody tr:hover {
        background: rgba(56,189,248,0.14) !important;
    }
    .skeleton-panel {
        background: rgba(17,25,40,0.85);
        border: 1px solid rgba(255,255,255,0.06);
        border-radius: 24px;
        padding: 2rem;
        margin: 1rem 0;
        box-shadow: 0 18px 50px rgba(0,0,0,0.3);
        animation: pulseGlow 1.8s ease-in-out infinite;
    }
    .skeleton-line {
        height: 16px;
        background: linear-gradient(90deg, rgba(255,255,255,0.08), rgba(255,255,255,0.18), rgba(255,255,255,0.08));
        border-radius: 999px;
        margin-bottom: 1rem;
    }
    .skeleton-line.short { width: 22%; }
    .skeleton-line.medium { width: 48%; }
    .skeleton-line.long { width: 86%; }
    .pulse-dot {
        display: inline-block;
        width: 12px;
        height: 12px;
        border-radius: 50%;
        margin-right: 0.6rem;
        background: rgba(56,189,248,0.7);
        box-shadow: 0 0 16px rgba(56,189,248,0.35);
        animation: pulseDot 1.4s ease-in-out infinite;
    }
    @keyframes pulseGlow {
        0%,100% { box-shadow: 0 18px 50px rgba(0,0,0,0.3); transform: translateY(0px); }
        50% { box-shadow: 0 24px 60px rgba(0,0,0,0.4); transform: translateY(-1px); }
    }
    @keyframes pulseDot {
        0%,100% { opacity: 0.4; transform: scale(1); }
        50% { opacity: 1; transform: scale(1.18); }
    }
    @keyframes fadeInSmooth {
        from { opacity: 0; transform: translateY(10px); }
        to { opacity: 1; transform: translateY(0); }
    }
    .custom-table-placeholder {
        min-height: 280px;
    }
    .stTextArea textarea {
        background: rgba(255,255,255,0.04) !important;
        color: #e5e7eb !important;
        border: 1px solid rgba(148, 163, 184, 0.18) !important;
    }
    .stTextArea label {
        color: #94a3b8 !important;
    }
    .stButton > button {
        background: linear-gradient(135deg, rgba(34,211,238,0.95), rgba(16,185,129,0.95));
        color: white;
        border-radius: 999px;
        padding: 0.8rem 1.35rem;
        font-weight: 600;
        border: none;
        transition: all 0.25s ease;
    }
    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 14px 30px rgba(34, 211, 238, 0.28);
    }
    </style>
    """, unsafe_allow_html=True)

def render_sidebar():
    """Sidebar config."""
    st.sidebar.markdown("""
    <div class="logo">
        <div class="network-node"></div>
        DATAX
    </div>
    """, unsafe_allow_html=True)
    
    st.sidebar.markdown('<div class="sidebar-glass metal-slider">', unsafe_allow_html=True)
    st.sidebar.markdown("⚙️ **الإعدادات**")
    st.sidebar.markdown("⏱️ الانتظار الذكي مفعل تلقائياً — يتم الانتظار حتى تصبح الشبكة خاملة وتظهر عناصر المحتوى الرئيسية.")
    st.sidebar.markdown('<div style="text-align: center; color: #00bcd4; font-weight: bold;">Smart waiting enabled</div>', unsafe_allow_html=True)
    st.sidebar.markdown('</div>', unsafe_allow_html=True)
    
    st.sidebar.markdown('<div class="sidebar-glass">', unsafe_allow_html=True)
    st.sidebar.markdown("📊 **الإحصائيات**")
    st.sidebar.markdown("""
    <div class="format-list">
        <div class="format-item">Excel</div>
        <div class="format-item">CSV</div>
        <div class="format-item">JSON</div>
        <div class="format-item">HTML</div>
        <div class="format-item">TXT</div>
    </div>
    """, unsafe_allow_html=True)
    st.sidebar.markdown('</div>', unsafe_allow_html=True)


def build_txt_from_entries(entries: List[dict]) -> str:
    rows = []
    for index, entry in enumerate(entries, start=1):
        rows.append(f"Entry {index}:")
        for key, value in entry.items():
            rows.append(f"{key}: {value}")
        rows.append('-' * 40)
    return '\n'.join(rows)


def filter_dataframe(df: pd.DataFrame, query: str) -> pd.DataFrame:
    if not query:
        return df
    mask = df.astype(str).apply(lambda row: row.str.contains(query, case=False, na=False), axis=1).any(axis=1)
    return df[mask]


def render_loading_placeholder(message: str = "جاري تحليل المحتوى...") -> str:
    return f"""
    <div class="skeleton-panel">
        <div class="pulse-dot"></div>
        <div style="font-size:1rem; color:#dbeafe; margin-bottom:1rem;">{message}</div>
        <div class="skeleton-line long"></div>
        <div class="skeleton-line medium"></div>
        <div class="skeleton-line short"></div>
        <div class="skeleton-line long"></div>
    </div>
    """


def execute_with_loader(action, message: str):
    placeholder = st.empty()
    placeholder.markdown(render_loading_placeholder(message), unsafe_allow_html=True)
    try:
        return action()
    finally:
        placeholder.empty()


def display_extracted_data(source_name: str, text_content: str, tables: List[pd.DataFrame], entries: List[dict]):
    """Professional results display."""
    st.markdown('<div class="glass-card result-panel">', unsafe_allow_html=True)
    st.markdown(f"""
        <div class="breadcrumb">
            <span>الرئيسية</span>
            <span>›</span>
            <span>نتائج الاستخراج</span>
            <span>›</span>
            <span>{source_name}</span>
        </div>
        <div class="toast-notice">✅ تمت عملية الاستخراج بنجاح</div>
    """, unsafe_allow_html=True)

    total_lines = len(text_content.splitlines()) if text_content else 0
    total_rows = sum(len(t) for t in tables)
    card_definitions = [
        ("سطور النص", total_lines, "metric-accent-cyan"),
        ("الجداول", len(tables), "metric-accent-emerald"),
        ("الكيانات", len(entries), "metric-accent-cyan"),
        ("إجمالي الصفوف", total_rows, "metric-accent-emerald"),
    ]
    cols = st.columns(4, gap="small")
    for col, (title, value, accent) in zip(cols, card_definitions):
        with col:
            st.markdown(f"""
            <div class="metric-card {accent}">
                <h3>{title}</h3>
                <div class="metric-value">{value}</div>
            </div>
            """, unsafe_allow_html=True)

    search_query = st.text_input("Search / Filter", value="", placeholder="ابحث في النتائج مباشرةً...", key="results_search", label_visibility="visible")
    action_col, spacer = st.columns([3,1], gap="small")
    with action_col:
        if entries:
            df_entries = pd.DataFrame(entries)
            filtered_entries = filter_dataframe(df_entries.fillna(''), search_query)
        else:
            filtered_entries = pd.DataFrame()

    st.markdown('<div class="action-bar">', unsafe_allow_html=True)
    st.markdown('<div></div>', unsafe_allow_html=True)
    st.markdown('<div class="action-toolbar">', unsafe_allow_html=True)
    if entries:
        csv_data = filtered_entries.to_csv(index=False, encoding='utf-8')
        json_data = json.dumps(filtered_entries.to_dict(orient='records'), ensure_ascii=False, indent=2)
        txt_data = build_txt_from_entries(filtered_entries.to_dict(orient='records'))
        st.download_button('⬇', data=csv_data.encode('utf-8'), file_name=f'{source_name}_filtered.csv', mime='text/csv', help='Download filtered CSV')
        st.download_button('📋', data=json_data.encode('utf-8'), file_name=f'{source_name}_filtered.json', mime='application/json', help='Download filtered JSON')
        st.download_button('👁️', data=txt_data.encode('utf-8'), file_name=f'{source_name}_filtered.txt', mime='text/plain', help='View filtered text snapshot')
    else:
        st.markdown('<span style="color:#94a3b8;">لا توجد نتائج للتحميل</span>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    tab1, tab2, tab3, tab4 = st.tabs(["🧾 الكيانات", "📄 النص الخام", "📊 الجداول", "📥 التصدير"])

    with tab1:
        if entries:
            st.markdown('<div class="dataframe-container">', unsafe_allow_html=True)
            st.dataframe(filtered_entries if not filtered_entries.empty else df_entries, use_container_width=True, hide_index=True)
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.info("لم يتم العثور على كيانات منظمة")

    with tab2:
        if text_content:
            st.markdown('<div class="glass-card">', unsafe_allow_html=True)
            st.text_area("", text_content, height=360)
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.info("لم يتم العثور على نص")

    with tab3:
        if tables:
            for i, df in enumerate(tables, 1):
                with st.expander(f"جدول {i} ({len(df)} صف × {len(df.columns)} عمود)"):
                    st.markdown('<div class="dataframe-container">', unsafe_allow_html=True)
                    st.dataframe(df, use_container_width=True, hide_index=True)
                    st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.info("لم يتم العثور على جداول")

    with tab4:
        col1, col2 = st.columns(2, gap="small")
        with col1:
            if entries and OPENPYXL_AVAILABLE:
                try:
                    df_entries = pd.DataFrame(entries)
                    excel_data = generate_entries_excel(source_name, df_entries)
                    st.download_button("⬇", excel_data, f"{source_name}_entries.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", help='Download entries Excel')
                except Exception as e:
                    st.error(f"Excel: {e}")
            if entries:
                txt_data = build_txt_from_entries(entries)
                st.download_button("📄", data=txt_data.encode('utf-8'), file_name=f"{source_name}_entries.txt", help='Download entries TXT')
            elif text_content:
                st.download_button("📄", data=text_content.encode('utf-8'), file_name=f"{source_name}.txt", help='Download extracted text TXT')

        with col2:
            if entries:
                df_entries = pd.DataFrame(entries)
                csv_data = df_entries.to_csv(index=False, encoding='utf-8')
                st.download_button("📋", data=csv_data.encode('utf-8'), file_name=f"{source_name}_entries.csv", help='Download entries CSV')
                json_data = json.dumps(entries, ensure_ascii=False, indent=2)
                st.download_button("🔗", data=json_data.encode('utf-8'), file_name=f"{source_name}_entries.json", help='Download entries JSON')
            else:
                if tables:
                    all_csv = "\n\n".join(df.to_csv(index=False, encoding='utf-8') for df in tables)
                    st.download_button("📋", data=all_csv.encode('utf-8'), file_name=f"{source_name}_tables.csv", help='Download tables CSV')
                json_data = {"source": source_name, "text": text_content, "tables_count": len(tables)}
                st.download_button("🔗", data=json.dumps(json_data, ensure_ascii=False, indent=2).encode('utf-8'), file_name=f"{source_name}.json", help='Download summary JSON')

    st.markdown('</div>', unsafe_allow_html=True)

# ============================================================================
# MAIN APP
# ============================================================================
def main():
    setup_page()
    
    # Title Card
    st.markdown("""
    <div class="title-card">
        <h1 style='font-size: 3rem; color: #ffffff; margin-bottom: 0.5rem;'>
            <span style='font-size: 2.5rem;'>🔍</span> استخراج البيانات المهني
        </h1>
        <p style='font-size: 1.2rem; color: #cccccc; margin-bottom: 0.5rem;'>استخراج النصوص والجداول من الروابط والملفات باحترافية كاملة</p>
        <div style='font-size: 1.1rem; color: #00bcd4;'>Excel • CSV • JSON • HTML • TXT | دعم كامل للعربية</div>
    </div>
    """, unsafe_allow_html=True)
    
    render_sidebar()
    
    # Action Panel
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    
    # Tabs inside the panel
    tab_url, tab_file = st.tabs(["🌐 رابط ويب", "📁 رفع ملف"])
    
    with tab_url:
        col1, col2 = st.columns([4,1])
        with col1:
            url = st.text_input("", placeholder="https://example.com", label_visibility="collapsed")
        with col2:
            st.markdown('<div style="padding-top: 0.5rem;"><span style="font-size: 1.5rem;">🔍</span></div>', unsafe_allow_html=True)
        
        # Output format dropdown
        output_format = st.selectbox("صيغة المخرج", ["Excel", "CSV", "JSON", "TXT"], index=0, label_visibility="visible")
        
        # Execution button
        if st.button("ابدأ عملية الاستخراج", type="primary"):
            if not validate_url(url):
                st.error("رابط غير صالح")
            else:
                try:
                    text, tables, entries = execute_with_loader(lambda: scrape_url(url), "جاري استخراج المحتوى والتحليل...")
                    display_extracted_data(url, text, tables, entries)
                except Exception as e:
                    st.error(f"خطأ: {e}")
    
    with tab_file:
        uploaded = st.file_uploader("اختر الملف", type=SUPPORTED_EXTENSIONS, label_visibility="visible")
        if uploaded:
            size = len(uploaded.getvalue())
            st.info(f"**{uploaded.name}** ({size/1024:.1f} KB)")
            if size > MAX_FILE_SIZE:
                st.error("الملف كبير جداً")
            else:
                output_format_file = st.selectbox("صيغة المخرج", ["Excel", "CSV", "JSON", "TXT"], index=0, key="file_format")
                if st.button("ابدأ عملية الاستخراج", type="primary", key="file_extract"):
                    try:
                        text, tables, entries = execute_with_loader(lambda: process_file(uploaded), "جاري معالجة الملف واستخراج البيانات...")
                        display_extracted_data(uploaded.name, text, tables, entries)
                    except Exception as e:
                        st.error(f"خطأ في المعالجة: {e}")
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # Status Bar Footer
    st.markdown("""
    <div class="status-bar">
        ✅ جاهز لاستخراج البيانات...
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()

